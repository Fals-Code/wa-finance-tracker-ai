import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

data = json.loads(sys.stdin.read())
rows = data['rows']
outpath = data['outpath']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Transaksi"


HEADER_BG   = "1E3A5F"   
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
KELUAR_BG   = "FFF2CC"   
MASUK_BG    = "E2EFDA"   
ALT_BG      = "F5F5F5"  
BORDER_COLOR = "CCCCCC"

thin = Side(style='thin', color=BORDER_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["No", "Tanggal", "Judul", "Nama Toko", "Nominal (Rp)", "Tipe", "Kategori", "Sub Kategori", "Catatan"]
col_widths = [5, 14, 28, 22, 16, 12, 20, 20, 30]


for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 30

for i, row in enumerate(rows, 2):
    tipe = (row.get("tipe") or "").lower()
    if tipe == "masuk":
        row_bg = MASUK_BG
    elif tipe == "keluar":
        row_bg = KELUAR_BG
    else:
        row_bg = ALT_BG if i % 2 == 0 else "FFFFFF"

    fill = PatternFill("solid", fgColor=row_bg)
    base_font = Font(name="Arial", size=10)

    vals = [
        i - 1,
        row.get("tanggal", ""),
        row.get("judul", ""),
        row.get("nama_toko", ""),
        row.get("nominal", 0),
        "💸 Keluar" if tipe == "keluar" else "💰 Masuk",
        row.get("kategori", ""),
        row.get("sub_kategori", ""),
        row.get("catatan", ""),
    ]

    for col, val in enumerate(vals, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.fill = fill
        cell.font = base_font
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 9))


        if col == 5:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 2:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 6:
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[i].height = 18

total_row = len(rows) + 2
ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="1E3A5F")
ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)

keluar_vals = [r.get("nominal", 0) for r in rows if (r.get("tipe") or "").lower() == "keluar"]
masuk_vals  = [r.get("nominal", 0) for r in rows if (r.get("tipe") or "").lower() == "masuk"]

tot_cell = ws.cell(row=total_row, column=5)
tot_cell.value = sum(keluar_vals) - sum(masuk_vals)  # net
tot_cell.number_format = '#,##0'
tot_cell.font = Font(name="Arial", bold=True, size=10)
tot_cell.alignment = Alignment(horizontal="right", vertical="center")
tot_cell.fill = PatternFill("solid", fgColor="D9E1F2")
tot_cell.border = border

label_cell = ws.cell(row=total_row, column=4, value="Net (Masuk - Keluar):")
label_cell.font = Font(name="Arial", bold=True, size=10, italic=True)
label_cell.alignment = Alignment(horizontal="right")
label_cell.fill = PatternFill("solid", fgColor="D9E1F2")

for col in [2, 3, 6, 7, 8, 9]:
    c = ws.cell(row=total_row, column=col)
    c.fill = PatternFill("solid", fgColor="D9E1F2")
    c.border = border

ws.freeze_panes = "A2"

ws.auto_filter.ref = f"A1:I{len(rows)+1}"

wb.save(outpath)
print("ok")